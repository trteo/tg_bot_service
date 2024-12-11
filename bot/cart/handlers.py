import json

from aiogram.exceptions import TelegramBadRequest
from aiogram.types import CallbackQuery, Message, LabeledPrice, FSInputFile, BufferedInputFile
from aiogram_dialog import StartMode, DialogManager
from aiogram_dialog.widgets.input import MessageInput
from aiogram_dialog.widgets.kbd import Button
from loguru import logger
from sqlalchemy import delete, update
from sqlalchemy import select

from bot.db.models import CartProducts, Product, Order, OrderProducts, OrderStatusEnum
from bot.db.session import async_session
from bot.states import CartStates
from bot.states import StartStates
from settings.config import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO


async def load_cart(dialog_manager: DialogManager, **kwargs):
    chat_id = dialog_manager.event.from_user.id

    async with async_session() as session:
        items_in_cart = (await
            session.execute(
                select(CartProducts, Product)
                .join(Product, CartProducts.product_id == Product.id)
                .filter(CartProducts.client_id == chat_id)
            )
        ).all()
    cart_items = [
        {
            "id": cart.id,
            "name": product.name,
            "quantity": cart.amount,
            "price": product.price,
            "total": product.price * cart.amount,
        }
        for cart, product in items_in_cart
    ]

    # Store the items in dialog data
    dialog_manager.dialog_data["cart_items"] = cart_items
    dialog_manager.dialog_data["index"] = 0


# Data getter to provide current cart items
async def get_cart_data(dialog_manager: DialogManager, **kwargs):
    if 'cart_items' not in dialog_manager.dialog_data:
        await load_cart(dialog_manager)

    cart_items = dialog_manager.dialog_data["cart_items"]
    if not cart_items:
        return {'cart_empty': True}

    index = dialog_manager.dialog_data.get("index", 0)
    item = cart_items[index]
    return {"item": item, "index": index + 1, "total_items": len(cart_items), 'cart_empty': False}


# Navigation callbacks handlers
async def on_next(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    cart_items = dialog_manager.dialog_data["cart_items"]
    if not cart_items:
        await dialog_manager.start(StartStates.MAIN, mode=StartMode.RESET_STACK)
        return

    current_index = dialog_manager.dialog_data.get("index", 0)
    new_index = (current_index + 1) % len(cart_items)
    dialog_manager.dialog_data["index"] = new_index
    await dialog_manager.switch_to(CartStates.VIEW_CART)


async def on_prev(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    cart_items = dialog_manager.dialog_data["cart_items"]
    if not cart_items:
        await dialog_manager.start(StartStates.MAIN, mode=StartMode.RESET_STACK)
        return

    current_index = dialog_manager.dialog_data.get("index", 0)
    new_index = (current_index - 1) % len(cart_items)
    dialog_manager.dialog_data["index"] = new_index
    await dialog_manager.switch_to(CartStates.VIEW_CART)


# Editing items
async def remove_item_from_cart_db(cart_product_id: int):
    async with async_session() as session:
        await session.execute(
            delete(CartProducts)
            .where(CartProducts.id == cart_product_id)
        )
        await session.commit()


async def set_item_in_cart_quantity_db(cart_product_id: int, quantity: int):
    async with async_session() as session:
        await session.execute(
            update(CartProducts)
            .where(CartProducts.id == cart_product_id)
            .values(amount=quantity)
        )
        await session.commit()


async def on_remove_item(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    cart_items = dialog_manager.dialog_data["cart_items"]
    if not cart_items:
        await dialog_manager.start(StartStates.MAIN, mode=StartMode.NEW_STACK)
        return
    current_index = dialog_manager.dialog_data.get("index", 0)
    await remove_item_from_cart_db(cart_product_id=int(cart_items.pop(current_index).get('id')))

    if cart_items:
        dialog_manager.dialog_data["index"] = current_index % len(cart_items)
    await dialog_manager.switch_to(CartStates.VIEW_CART)


async def on_set_quantity(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    await callback.message.answer("Send me the new quantity for this item.")
    dialog_manager.dialog_data["waiting_for_quantity"] = True


async def receive_quantity(message: Message, dialog_manager: DialogManager):
    dialog_manager.dialog_data["waiting_for_quantity"] = False

    new_quantity = int(message.text)
    cart_items = dialog_manager.dialog_data["cart_items"]
    current_index = dialog_manager.dialog_data.get("index", 0)

    if new_quantity <= 0:
        removed_item = cart_items.pop(current_index)
        await remove_item_from_cart_db(cart_product_id=removed_item.get('id'))
        await message.answer("Item removed from the cart due to zero or negative quantity.")
    else:
        cart_items[current_index]["quantity"] = new_quantity
        cart_items[current_index]["total"] = cart_items[current_index]["price"] * new_quantity
        await set_item_in_cart_quantity_db(
            cart_product_id=cart_items[current_index].get('id'),
            quantity=new_quantity
        )
        await message.answer(f"Quantity updated to {new_quantity}.")

    await dialog_manager.start(CartStates.VIEW_CART)


async def accept_delivery_address(message: Message, message_input: MessageInput, dialog_manager: DialogManager):
    address = message.text
    logger.info(f'Address: {address}')
    dialog_manager.current_context().dialog_data["delivery_address"] = address
    # TODO create order
    await dialog_manager.switch_to(CartStates.ADDRESS_CONFIRM)


async def get_entered_addr(dialog_manager: DialogManager, **kwargs):
    return {'entered_addr': dialog_manager.current_context().dialog_data["delivery_address"]}


# Making order
async def register_order(callback: CallbackQuery, button: Button, dialog_manager: DialogManager):
    user_id = dialog_manager.event.from_user.id
    delivery_address = dialog_manager.current_context().dialog_data["delivery_address"]

    # Create Order
    async with async_session() as session:
        order = Order(
            delivery_address=delivery_address,
            client_id=user_id,
            status=OrderStatusEnum.REGISTERED.value,
        )
        session.add(order)
        await session.commit()

    # Create Order Items
    async with async_session() as session:
        cart_items = (await session.execute(
            select(CartProducts, Product)
            .join(Product, CartProducts.product_id == Product.id)
            .filter(CartProducts.client_id == user_id)
        )).all()
        order_products = [
            OrderProducts(
                order_id=order.id,
                product_id=cart_item.product_id,
                amount=cart_item.amount,
                price=product.price,
            )
            for cart_item, product in cart_items
        ]

        session.add_all(order_products)
        await session.commit()

    # Flush users cart
        for cart_item, _ in cart_items:
            await session.delete(cart_item)
        await session.commit()
    dialog_manager.dialog_data['created_order_id'] = order.id
    await confirm_cart_and_send_invoice(
        event=dialog_manager.event,
        order_id=order.id,
        dialog_manager=dialog_manager
    )


async def confirm_cart_and_send_invoice(
        event: CallbackQuery,
        order_id: int,
        dialog_manager: DialogManager,
):
    logger.info(f'Created payment for order with id {order_id}')
    chat_id = event.from_user.id
    async with async_session() as session:
        order_items = (await session.execute(
            select(OrderProducts, Product)
            .join(Product, OrderProducts.product_id == Product.id)
            .filter(OrderProducts.order_id == order_id)
        )).all()

    # Build cart description and total cost
    description = "Your cart items:\n"
    prices = []
    total_price = 0
    for ordered_item, item in order_items:
        item_sun_price = float(item.price) * ordered_item.amount
        description += f"{item.name} (x{ordered_item.amount}): {item_sun_price} RUB\n"
        prices.append(LabeledPrice(label=item.name, amount=int(item_sun_price * 100)))  # Telegram uses cents
        total_price += item_sun_price

    description += f"\nTotal: {total_price} RUB"
    logger.info(f'prices: {prices}')
    try:
        await event.bot.send_invoice(
            chat_id=chat_id,
            title="Your Cart Purchase",
            description=description,
            payload=json.dumps({'order_id': order_id, 'status': 'success'}),
            provider_token=settings.PAYMENT_TOKEN.get_secret_value(),
            currency="RUB",
            prices=prices,
            start_parameter="purchase",
        )
    except TelegramBadRequest:
        await event.message.answer("Something went wrong. Try again later")
        await dialog_manager.switch_to(CartStates.VIEW_CART)


async def __generate_excel_report(order_id: int) -> BytesIO:
    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Summary"

    # Define headers
    headers = ["Name", "Quantity", "Price per Item", "Total Price"]
    ws.append(headers)  # Add headers to the first row

    # Style headers (bold and centered)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Add data rows
    total_cost = 0

    async with async_session() as session:
        ordered_products = (await session.execute(
            select(OrderProducts, Product)
            .filter(OrderProducts.order_id == order_id)
            .join(Product, OrderProducts.product_id == Product.id)
        )).all()

        for ordered_item, item_info in ordered_products:
            total_price = ordered_item.amount * item_info.price
            ws.append([item_info.name, ordered_item.amount, item_info.price, total_price])
            total_cost += total_price

    # Append the total cost row
    ws.append([])
    ws.append(["", "", "Total Order Cost", total_cost])
    total_row = ws.max_row

    # Style the total cost row (bold and centered)
    for col in ws.iter_cols(min_row=total_row, max_row=total_row, min_col=3, max_col=4):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Save the workbook to an in-memory bytes buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Move the pointer to the start of the file
    return buffer


async def handle_payment(message: Message, dialog_manager: DialogManager):
    payload = json.loads(message.successful_payment.invoice_payload)

    if payload.get('status', '') != "success":
        new_status = OrderStatusEnum.PAY_FAILED
        await message.answer("Sorry, something went wrong")
    else:
        new_status = OrderStatusEnum.PAID
        await message.answer(
            "Thank you! Your payment was successful. Your order is now being processed"
        )

        excel_buffer = await __generate_excel_report(order_id=payload.get('order_id'))
        file = BufferedInputFile(excel_buffer.read(), filename="order_summary.xlsx")
        await message.answer_document(file, caption="Here is your order summary")

    async with async_session() as session:
        await session.execute(
            update(Order)
            .where(Order.id == payload.get('order_id'))
            .values(status=new_status.value)
        )
        await session.commit()

    await dialog_manager.start(StartStates.MAIN, mode=StartMode.RESET_STACK)